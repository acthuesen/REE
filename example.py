#Depending on setup, it's useful to do this for many files at a time. 
#This is set up to move files from a folder within the directory called 'new' to a folder within the directory called 'calc'. I.e. during data collection new files are placed in 'new' folder and calculated as required. 
#Results are added to a results file which can be used for downstream statistical procedures eg. 
#In this example, measurements are repeated and the measurement sequence is added to the filename during data export from the device. This replaces time. 

import pandas as pd
import os
import glob
import openpyxl as op
import shutil
import sys
sys._enablelegacywindowsfsencoding()

d = 'YOURPATH'
files = glob.glob(os.path.join(d, '*new\\', '*.xlsx'))

for file in files:
	full_data = pd.read_excel(file)
	#Extract ID and date, sequence (from filename), and height and weight, store as values
	ID = full_data.columns[1]
	date = full_data.columns[4]
	wt = full_data.iloc[5,1]
	ht = full_data.iloc[4,1]
	fnspl = file.split('_')
	fnspl2 = fnspl[-1].split('.')
	sequence = fnspl2[0]
	bmi = wt/(ht/100)**2
	#drop cols 0-8/select cols between t and VO2%Pred; drop rows 1-2
	subset_data = full_data.loc[:, 't':'VO2%Pred']
	cdat = subset_data.drop([0,1])
	#Infer dtypes
	cdat = cdat.infer_objects()
	mask = cdat['Phase']=='REST'
	restcdat = cdat[mask]
	mvo2 = restcdat['VO2'].mean()
	mvco2 = restcdat['VCO2'].mean()
	mrq = restcdat['RQ'].mean()
	meekc = restcdat['EEkc'].mean()
	meeh = restcdat['EEh'].mean()
	meem = restcdat['EEm'].mean()
	meetot = restcdat['EEtot'].mean()
	meekg = restcdat['EEkg'].mean()
	mpro = restcdat['PRO'].mean()
	mfat = restcdat['FAT'].mean()
	mcho = restcdat['CHO'].mean()
	#write results
	rfn = os.path.join(d, 'calorimetry_results.xlsx')
	wb = op.load_workbook(rfn)
	ws = wb.active
	maxr = ws.max_row + 1
	ws.cell(row = maxr, column = 1).value = ID
	ws.cell(row = maxr, column = 2).value = ht
	ws.cell(row = maxr, column = 3).value = wt
	ws.cell(row = maxr, column = 4).value = bmi
	ws.cell(row = maxr, column = 5).value = date
	ws.cell(row = maxr, column = 6).value = sequence
	ws.cell(row = maxr, column = 7).value = mvo2
	ws.cell(row = maxr, column = 8).value = mvco2
	ws.cell(row = maxr, column = 9).value = mrq
	ws.cell(row = maxr, column = 10).value = meekc
	ws.cell(row = maxr, column = 11).value = meeh
	ws.cell(row = maxr, column = 12).value = meem
	ws.cell(row = maxr, column = 13).value = meetot
	ws.cell(row = maxr, column = 14).value = meekg
	ws.cell(row = maxr, column = 15).value = mpro
	ws.cell(row = maxr, column = 16).value = mfat
	ws.cell(row = maxr, column = 17).value = mcho
	ws.cell(row = maxr, column = 18).value = file
	wb.save(rfn)
af = glob.glob(os.path.join(d, '*new\\*'))
source = os.path.join(d, 'new')
dest = os.path.join(d, 'calc')

for f in af:
	shutil.move(f,dest)
